import re
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# ====== CONFIG (your original paths preserved) ======
# Path to folder with .out files
OUT_DIR = Path("/scratch/users/t02pa24/Project MSC/Pytrich/logs/out")
# Path to save Excel output
OUTPUT_EXCEL = Path("/scratch/users/t02pa24/Project MSC/Pytrich/logtools/Exports/final_structured_output.xlsx")

PER_PROBLEM_CAP_SEC = 600
HEURISTICS = ["DELRELAX", "TDG", "HMAX"]
SUBCOLUMNS = [
    "Search Time",
    "Nodes Expanded",
    "Search Status",
    "Solution Size",
    "Revisits Avoided",
    "Fringe Size",
    "Memory Used (%)"
]
# ====================================================

# Regex patterns
heuristic_header_re = re.compile(r"=== Running heuristic: ([A-Z0-9_]+)")
solving_re = re.compile(r">>> Solving: ([^\s]+) \(cap (\d+)s\)")
search_status_re = re.compile(r"Search Status\s*:\s*([A-Z_]+)")
search_time_re = re.compile(r"Search Time \(seconds\)\s*:\s*([0-9.]+)")
nodes_expanded_re = re.compile(r"Nodes Expanded\s*:\s*([0-9]+)")
solution_size_re = re.compile(r"Solution Size\s*:\s*([0-9]+)")
revisits_re = re.compile(r"Revisits Avoided:\s*([0-9]+)")
fringe_size_re = re.compile(r"Fringe size\s*:\s*([0-9]+)")
used_mem_re = re.compile(r"Used Memory:\s*([0-9.]+)%")
failed_rc_re = re.compile(r"Result:\s*FAILED\(rc=(\d+)\)")
invalid_graph_re = re.compile(r"Invalid Graph Type\s*(\d+)")

def parse_out_file(file_path):
    rows = []
    domain_name = Path(file_path).stem.replace("out_domain_", "")
    text = Path(file_path).read_text(errors="ignore")
    lines = text.splitlines()

    current_heuristic = None
    i = 0
    while i < len(lines):
        m_h = heuristic_header_re.search(lines[i])
        if m_h:
            current_heuristic = m_h.group(1).upper()
            i += 1
            continue

        m_s = solving_re.search(lines[i])
        if m_s:
            problem = m_s.group(1)
            cap = int(m_s.group(2))
            status = None
            time_s = None
            nodes_exp = None
            sol_size = None
            revisits = None
            fringe = None
            used_mem = None
            invalid_graph = None
            failed_rc = None

            j = i + 1
            while j < len(lines):
                if solving_re.search(lines[j]) or heuristic_header_re.search(lines[j]):
                    break
                if status is None:
                    m_status = search_status_re.search(lines[j])
                    if m_status:
                        status = m_status.group(1)
                if time_s is None:
                    m_time = search_time_re.search(lines[j])
                    if m_time:
                        time_s = float(m_time.group(1))
                if nodes_exp is None:
                    m_ne = nodes_expanded_re.search(lines[j])
                    if m_ne:
                        nodes_exp = int(m_ne.group(1))
                if sol_size is None:
                    m_ss = solution_size_re.search(lines[j])
                    if m_ss:
                        sol_size = int(m_ss.group(1))
                if revisits is None:
                    m_rv = revisits_re.search(lines[j])
                    if m_rv:
                        revisits = int(m_rv.group(1))
                if fringe is None:
                    m_fs = fringe_size_re.search(lines[j])
                    if m_fs:
                        fringe = int(m_fs.group(1))
                if used_mem is None:
                    m_um = used_mem_re.search(lines[j])
                    if m_um:
                        used_mem = float(m_um.group(1))
                if invalid_graph is None:
                    m_ig = invalid_graph_re.search(lines[j])
                    if m_ig:
                        invalid_graph = int(m_ig.group(1))
                if failed_rc is None:
                    m_fail = failed_rc_re.search(lines[j])
                    if m_fail:
                        failed_rc = int(m_fail.group(1))
                j += 1

            # Canonical status
            if invalid_graph is not None:
                canonical = "ERROR_INVALID_GRAPH"
                time_s = None
            elif failed_rc == 143:
                canonical = "TIMEOUT_OR_KILLED"
                if time_s is None:
                    time_s = cap
            elif status == "GOAL":
                canonical = "SOLVED"
            elif status == "UNSOLVABLE":
                canonical = "UNSOLVABLE"
            else:
                canonical = "UNKNOWN"

            rows.append({
                "domain": domain_name,
                "problem": problem,
                "heuristic": current_heuristic,
                "Search Time": time_s,
                "Nodes Expanded": nodes_exp,
                "Search Status": canonical,
                "Solution Size": sol_size,
                "Revisits Avoided": revisits,
                "Fringe Size": fringe,
                "Memory Used (%)": used_mem
            })
            i = j
            continue
        i += 1
    return rows

def main():
    all_rows = []
    for file in sorted(OUT_DIR.glob("*.out")):
        all_rows.extend(parse_out_file(file))

    df = pd.DataFrame(all_rows)

    # Pivot to wide format
    final_rows = []
    for (domain, problem), group in df.groupby(["domain", "problem"]):
        row_data = {"Domain": domain, "Problem": problem}
        for h in HEURISTICS:
            h_data = group[group["heuristic"] == h]
            if not h_data.empty:
                r = h_data.iloc[0]
                for col in SUBCOLUMNS:
                    row_data[(h, col)] = r[col]
            else:
                for col in SUBCOLUMNS:
                    row_data[(h, col)] = None
        final_rows.append(row_data)

    final_df = pd.DataFrame(final_rows)
    final_df.sort_values(["Domain", "Problem"], inplace=True)

    # Write to Excel with merged headers and vertical merged domains
    wb = Workbook()
    ws = wb.active

    # First header row
    ws.cell(row=1, column=1, value="Domain")
    ws.merge_cells(start_row=1, end_row=2, start_column=1, end_column=1)
    ws.cell(row=1, column=2, value="Problem")
    ws.merge_cells(start_row=1, end_row=2, start_column=2, end_column=2)

    col_idx = 3
    for h in HEURISTICS:
        ws.cell(row=1, column=col_idx, value=h)
        ws.merge_cells(start_row=1, end_row=1, start_column=col_idx, end_column=col_idx + len(SUBCOLUMNS) - 1)
        for j, subcol in enumerate(SUBCOLUMNS):
            ws.cell(row=2, column=col_idx + j, value=subcol)
        col_idx += len(SUBCOLUMNS)

    # Fill data rows
    for i, row in enumerate(final_df.itertuples(index=False), start=3):
        ws.cell(row=i, column=1, value=row[0])  # Domain
        ws.cell(row=i, column=2, value=row[1])  # Problem
        c_idx = 2
        for val in row[2:]:
            c_idx += 1
            ws.cell(row=i, column=c_idx, value=val)

    # Merge domain cells vertically
    start_row = 3
    for i in range(3, len(final_df) + 3):
        current_domain = ws.cell(row=i, column=1).value
        next_domain = ws.cell(row=i + 1, column=1).value if i < len(final_df) + 2 else None
        if current_domain != next_domain:
            if start_row != i:
                ws.merge_cells(start_row=start_row, end_row=i, start_column=1, end_column=1)
                ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            start_row = i + 1

    # Styling for headers
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)

    wb.save(OUTPUT_EXCEL)
    print(f"Saved Excel to {OUTPUT_EXCEL}")

if __name__ == "__main__":
    main()