import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import sys

# ====== CONFIG ======
INPUT_FILE = Path("/scratch/users/t02pa24/Project MSC/Pytrich/logtools/Exports/final_structured_output.xlsx")
OUTPUT_EXCEL = Path("/scratch/users/t02pa24/Project MSC/Pytrich/logtools/Exports/summary_grouped.xlsx")

HEURISTICS = ["DELRELAX", "TDG", "HMAX"]
SUBCOLUMNS = [
    "Coverage (%)",
    "Avg Search Time",
    "Avg Nodes Expanded",
    "Avg Solution Size",
    "Total Problems",
    "Solved Problems",
]
# ====================

def main():
    if not INPUT_FILE.exists():
        sys.exit(f"[ERROR] Input file not found: {INPUT_FILE}")

    # Read with two header rows
    df = pd.read_excel(INPUT_FILE, header=[0, 1])
    print(f"[INFO] Loaded {df.shape[0]} rows and {df.shape[1]} columns from {INPUT_FILE}")

    if df.empty:
        sys.exit("[ERROR] The input file is empty.")

    # Flatten headers
    df.columns = [
        " ".join([str(c).strip() for c in col if pd.notna(c) and str(c).strip()])
        for col in df.columns.values
    ]

    # Fix first two columns
    df.rename(columns={df.columns[0]: "Domain", df.columns[1]: "Problem"}, inplace=True)

    # Fill missing Domain values (merged cells in original file)
    df["Domain"] = df["Domain"].ffill()

    results = []

    # Group by Domain
    for domain, domain_df in df.groupby("Domain"):
        row_data = {"Domain": domain}
        total_problems = len(domain_df)

        for heuristic in HEURISTICS:
            required_cols = [
                f"{heuristic} Search Time",
                f"{heuristic} Nodes Expanded",
                f"{heuristic} Search Status",
                f"{heuristic} Solution Size",
            ]
            for col in required_cols:
                if col not in domain_df.columns:
                    print(f"[WARNING] Missing column: {col} for domain {domain}")
                    domain_df[col] = None

            h_df = domain_df[required_cols]

            solved_mask = (
                h_df[f"{heuristic} Search Status"]
                .astype(str)
                .str.strip()
                .str.upper()
                .isin(["SOLVED", "GOAL"])
            )

            solved_count = solved_mask.sum()
            coverage = (solved_count / total_problems) * 100 if total_problems > 0 else 0

            if solved_count > 0:
                avg_time = h_df.loc[solved_mask, f"{heuristic} Search Time"].mean()
                avg_nodes = h_df.loc[solved_mask, f"{heuristic} Nodes Expanded"].mean()
                avg_sol_size = h_df.loc[solved_mask, f"{heuristic} Solution Size"].mean()

                row_data[f"{heuristic} Coverage (%)"] = round(coverage, 2)
                row_data[f"{heuristic} Avg Search Time"] = round(avg_time, 2)
                row_data[f"{heuristic} Avg Nodes Expanded"] = round(avg_nodes, 2)
                row_data[f"{heuristic} Avg Solution Size"] = round(avg_sol_size, 2)
            else:
                row_data[f"{heuristic} Coverage (%)"] = 0
                row_data[f"{heuristic} Avg Search Time"] = "UNSOLVED"
                row_data[f"{heuristic} Avg Nodes Expanded"] = "UNSOLVED"
                row_data[f"{heuristic} Avg Solution Size"] = "UNSOLVED"

            row_data[f"{heuristic} Total Problems"] = total_problems
            row_data[f"{heuristic} Solved Problems"] = solved_count

        results.append(row_data)

    final_df = pd.DataFrame(results)
    final_df.sort_values(["Domain"], inplace=True)

    # Write to Excel
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Domain")
    ws.merge_cells(start_row=1, end_row=2, start_column=1, end_column=1)

    col_idx = 2
    for h in HEURISTICS:
        ws.cell(row=1, column=col_idx, value=h)
        ws.merge_cells(start_row=1, end_row=1, start_column=col_idx, end_column=col_idx + len(SUBCOLUMNS) - 1)
        for j, subcol in enumerate(SUBCOLUMNS):
            ws.cell(row=2, column=col_idx + j, value=subcol)
        col_idx += len(SUBCOLUMNS)

    # ? Fixed loop to write exact matching column names
    for i in range(len(final_df)):
        ws.cell(row=i+3, column=1, value=final_df.iloc[i]["Domain"])
        col_idx = 2
        for h in HEURISTICS:
            for subcol in SUBCOLUMNS:
                col_name = f"{h} {subcol}"
                if col_name in final_df.columns:
                    val = final_df.iloc[i][col_name]
                else:
                    val = None
                ws.cell(row=i+3, column=col_idx, value=val)
                col_idx += 1

    # Merge domain cells vertically
    start_row = 3
    for i in range(3, len(final_df) + 3):
        current_domain = ws.cell(row=i, column=1).value
        next_domain = ws.cell(row=i + 1, column=1).value if i < len(final_df) + 2 else None
        if current_domain != next_domain:
            if start_row != i:
                ws.merge_cells(start_row=start_row, end_row=i, start_column=1, end_column=1)
                ws.cell(start_row, 1).alignment = Alignment(horizontal="center", vertical="center")
            start_row = i + 1

    # Format headers
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row[1:]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_idx in range(2, ws.max_column + 1, len(SUBCOLUMNS)):
            ws.cell(row=row[0].row, column=col_idx).font = Font(bold=True)

    wb.save(OUTPUT_EXCEL)
    print(f"[SUCCESS] Saved grouped summary to {OUTPUT_EXCEL}")

if __name__ == "__main__":
    main()
