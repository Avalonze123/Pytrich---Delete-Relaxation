#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import csv
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd


def parse_args():
    p = argparse.ArgumentParser(
        description="Summarize parsed logs by Domain with heuristic column-groups "
                    "(coverage + averages), formatted like the grouped Excel output your professor liked."
    )
    p.add_argument("-i", "--input", required=True,
                   help="Input CSV/XLSX from parse_log.py (e.g., parsed_log_summary.csv or .xlsx)")
    p.add_argument("-o", "--output_xlsx", default="summary_grouped.xlsx",
                   help="Output Excel filename (default: summary_grouped.xlsx)")
    p.add_argument("--csv", dest="output_csv", default=None,
                   help="Optional: also write a flat summary CSV")
    p.add_argument("--heuristic", default=None,
                   help="Only include rows matching this heuristic (exact match)")
    return p.parse_args()


def safe_float(x, default=0.0):
    try:
        return float(x)
    except:
        return default


def safe_int(x, default=0):
    try:
        return int(float(x))
    except:
        return default


def load_data(input_file):
    """Loads CSV or XLSX into a list of dict rows."""
    ext = os.path.splitext(input_file)[1].lower()
    if ext == ".csv":
        with open(input_file, "r", encoding="utf-8") as f:
            return list(csv.DictReader(f))
    elif ext in (".xlsx", ".xls"):
        df = pd.read_excel(input_file)
        df = df.fillna("")  # replace NaN with empty string
        return df.to_dict(orient="records")
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def aggregate(input_file, heuristic_filter=None):
    stats = defaultdict(lambda: {
        "total": 0,
        "solved": 0,
        "sum_time": 0.0,
        "sum_nodes": 0,
        "sum_size": 0,
    })

    domains = set()
    heuristics = set()

    rows = load_data(input_file)

    for row in rows:
        domain = (row.get("Domain") or "").strip()
        heuristic = (row.get("Heuristic") or "").strip()
        status = (row.get("Search Status") or row.get("Status") or "").strip().upper()
        time_s = safe_float(row.get("Search Time", "0"))
        nodes = safe_int(row.get("Nodes Expanded", "0"))
        size = safe_int(row.get("Solution Size", "0"))

        if not domain or not heuristic:
            continue

        if heuristic_filter and heuristic != heuristic_filter:
            continue

        key = (domain, heuristic)
        stats[key]["total"] += 1
        stats[key]["sum_time"] += time_s
        stats[key]["sum_nodes"] += nodes
        stats[key]["sum_size"] += size

        if status == "GOAL":
            stats[key]["solved"] += 1

        domains.add(domain)
        heuristics.add(heuristic)

    return stats, sorted(domains), sorted(heuristics)


def compute_row(stats, domain, heuristic):
    d = stats.get((domain, heuristic), None)
    if not d:
        return ["", "", "", "", "", ""]

    total = d["total"]
    solved = d["solved"]
    coverage = (solved / total * 100.0) if total else 0.0

    # Average over ALL problems (solved + unsolved)
    if total > 0:
        avg_time = d["sum_time"] / total
        avg_nodes = d["sum_nodes"] / total
        avg_size = d["sum_size"] / total
    else:
        avg_time = 0.0
        avg_nodes = 0.0
        avg_size = 0.0

    return [
        f"{coverage:.2f}",
        f"{avg_time:.3f}",
        f"{avg_nodes:.2f}",
        f"{avg_size:.2f}",
        str(total),
        str(solved),
    ]


def write_excel(output_xlsx, stats, domains, heuristics):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    subcols = ["Coverage (%)", "Avg Search Time (s)", "Avg Nodes Expanded", "Avg Solution Size", "Total Problems", "Solved"]

    header_top = ["Domain"]
    header_sub = [""]

    for h in heuristics:
        header_top += [h] + [""] * (len(subcols) - 1)
        header_sub += subcols

    ws.append(header_top)
    ws.append(header_sub)

    start_col = 2
    for _h in heuristics:
        end_col = start_col + len(subcols) - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        start_col = end_col + 1

    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in (1, 2):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

    row_idx = 3
    for domain in domains:
        row = [domain]
        for h in heuristics:
            row.extend(compute_row(stats, domain, h))
        ws.append(row)
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
        row_idx += 1

    ws.column_dimensions["A"].width = 26
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    for c in range(1, ws.max_column + 1):
        ws.cell(row=1, column=c).alignment = center
        ws.cell(row=2, column=c).alignment = center

    wb.save(output_xlsx)


def write_flat_csv(output_csv, stats, domains, heuristics):
    with open(output_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "Domain", "Heuristic",
            "Total Problems", "Solved", "Coverage (%)",
            "Avg Search Time (s)", "Avg Nodes Expanded", "Avg Solution Size"
        ])
        for domain in domains:
            for h in heuristics:
                vals = compute_row(stats, domain, h)
                if vals == ["", "", "", "", "", ""]:
                    continue
                cov, avg_t, avg_n, avg_s, total, solved = vals
                w.writerow([domain, h, total, solved, cov, avg_t, avg_n, avg_s])


def main():
    args = parse_args()
    if not os.path.exists(args.input):
        print(f"[ERROR] Input file not found: {args.input}")
        return

    stats, domains, heuristics = aggregate(args.input, heuristic_filter=args.heuristic)

    write_excel(args.output_xlsx, stats, domains, heuristics)
    print(f"[OK] Wrote grouped Excel: {args.output_xlsx}")

    if args.output_csv:
        write_flat_csv(args.output_csv, stats, domains, heuristics)
        print(f"[OK] Wrote flat CSV: {args.output_csv}")


if __name__ == "__main__":
    main()
